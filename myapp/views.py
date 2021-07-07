from django.shortcuts import render
import openpyxl
import xlwt
from xlwt import Workbook
from geopy.geocoders import Nominatim


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()

        wb = Workbook()
  
        # add_sheet is used to create sheet.
        sheet1 = wb.add_sheet('Sheet 1')
        style = xlwt.easyxf('font: bold 1, color blue;')
        sheet1.write(0, 0, 'Address', style)
        sheet1.write(0, 1, 'Latitude', style)
        sheet1.write(0, 2, 'Longitude', style)
        wb.save('xlwt-solution.xls')
        geolocator = Nominatim(user_agent="my_user_agent")
        # iterating over the rows and
        # getting value from each cell in row
        i = 1
        j = 0
        for row in worksheet.iter_cols(min_col=1, max_col=1):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
                address = cell.value
                loc = geolocator.geocode(address)
                print("latitude is :-" ,loc.latitude,"\nlongtitude is:-" ,loc.longitude)
                row_data.append(str(loc.latitude))
                row_data.append(str(loc.longitude))
                sheet1.write(i, j, str(cell.value))
                sheet1.write(i, j+1, str(loc.latitude))
                sheet1.write(i, j+2, str(loc.longitude))
                i = i+1
                wb.save('xlwt-solution.xls')
                #row_data.append("\n")
                #rows_data = '\n'.join(row_data)
            excel_data.append("")

        return render(request, 'myapp/index.html', {"excel_data":excel_data})








